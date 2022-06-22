import openpyxl
from openpyxl.styles import PatternFill  # Подключаем стили для ячеек
from openpyxl.styles import Font  # Подключаем стили для текста
from openpyxl.styles import colors  # Подключаем цвета для текста и ячеек


def write_report(epics_list: list, tasks_list: list, subtasks_list: list, date_str: str):
    workbook = openpyxl.Workbook()
    worksheet = workbook['Sheet']

    worksheet['A1'] = f'Отчет по просроченным задачам команды за период ({date_str})'
    title_font(worksheet['A1'])
    worksheet['A2'] = 'Проект'
    worksheet['B2'] = 'Эпик'
    worksheet['C2'] = 'Задача'
    worksheet['D2'] = 'Подзадача'
    worksheet['E2'] = 'Исполнитель'
    worksheet['F2'] = 'Статус'
    worksheet['G2'] = 'Описание'
    worksheet['H2'] = 'Дата начала'
    worksheet['I2'] = 'Дата завершения'
    worksheet['J2'] = 'Фактическая дата завершения'
    header_style(workbook)
    column_num = 3
    for epic in epics_list:
        for task in tasks_list:
            i = str(column_num)
            if epic.key == task.epic.key:
                worksheet['A' + i] = epic.project
                worksheet['B' + i] = epic.issue_key
                worksheet['C' + i] = task.issue_key
                worksheet['E' + i] = task.assignee
                worksheet['F' + i] = task.status
                worksheet['G' + i] = task.summary
                worksheet['H' + i] = task.start_date
                worksheet['I' + i] = task.due_date
                worksheet['J' + i] = task.real_date
                column_num += 1
                for subtask in subtasks_list:
                    i = str(column_num)
                    if task.key == subtask.task:
                        worksheet['D' + i] = subtask.issue_key
                        worksheet['E' + i] = subtask.assignee
                        worksheet['F' + i] = subtask.status
                        worksheet['G' + i] = subtask.summary
                        worksheet['H' + i] = subtask.start_date
                        worksheet['I' + i] = subtask.due_date
                        worksheet['J' + i] = subtask.real_date
                        column_num += 1
    excel_filename = date_str + '_report_overdue.xlsx'
    workbook.save(excel_filename)
    print('Отчет готов')

    return excel_filename

def title_font(work_sheet):
    work_sheet_buf = work_sheet
    work_sheet_buf.font = Font(size=18, underline='single', color='000080', bold=True, italic=True)
# текст: размер — 23, подчеркивание, цвет = FFBB00 можно (color=colors.RED), жирный, наклонный.

def header_style(wb):
    ws = wb.active
    fill = PatternFill("solid", fgColor="1E90FF")
    for cell in list(ws.rows)[1]:
        cell.fill = fill
# Цвет ячейки