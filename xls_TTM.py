import openpyxl
from openpyxl.styles import PatternFill  # Подключаем стили для ячеек
from openpyxl.styles import Font  # Подключаем стили для текста
from openpyxl.styles import colors  # Подключаем цвета для текста и ячеек


def write_report(data: list, date_str: str):
    workbook = openpyxl.Workbook()
    worksheet = workbook['Sheet']

    worksheet['A1'] = f'Отчет по эпикам, которые превысили Т2М в 90 дней'
    title_font(worksheet['A1'])
    worksheet['A2'] = 'Проект'
    worksheet['B2'] = 'Ключ задачи'
    worksheet['C2'] = 'Исполнитель'
    worksheet['D2'] = 'Тип'
    worksheet['E2'] = 'Статус'
    worksheet['F2'] = 'Описание'
    worksheet['G2'] = 'Дата начала'
    worksheet['H2'] = 'Дата завершения'
    worksheet['I2'] = 'Фактическая дата завершения'
    header_style(workbook)
    column_num = 3
    for item in data:
        i = str(column_num)
        worksheet['A' + i] = item.project_name
        worksheet['B' + i] = item.issue_key
        worksheet['C' + i] = item.assignee
        worksheet['D' + i] = item.issue_type
        worksheet['E' + i] = item.status
        worksheet['F' + i] = item.summary
        worksheet['G' + i] = item.start_date
        worksheet['H' + i] = item.due_date
        worksheet['I' + i] = item.real_date
        column_num += 1
    excel_filename = date_str + '_report_TTM.xlsx'
    workbook.save(excel_filename)
    print('Отчет готов')

    return excel_filename


def title_font(work_sheet):
    work_sheet_buf = work_sheet
    work_sheet_buf.font = Font(size=18, underline='single', color='000080', bold=True, italic=True)
# текст: размер — 23, подчеркивание, цвет = FFBB00 можно (color=colors.RED), жирный, наклонный.

def header_style(wb):
    ws = wb.active
    fill = PatternFill("solid", fgColor="1E90FF")
    for cell in list(ws.rows)[1]:
        cell.fill = fill
# Цвет ячейки